"""
app/services/import_service.py

Import service: CSV/Excel file parsing, column auto-mapping via difflib,
Pydantic-based row validation, duplicate detection, and error CSV generation.

Security notes:
- All cell values from user uploads treated as untrusted.
- openpyxl loaded with read_only=True, data_only=True to prevent formula evaluation
  and write-back attacks (T-04-01).
- csv.DictReader handles quoting/escaping; utf-8-sig decoding strips BOM (T-04-02).
- All cell values cast to str to prevent injection (T-04-02).
"""
from __future__ import annotations

import csv
import io
from difflib import SequenceMatcher
from typing import Any

import openpyxl

from app.platform.yaml_loader import EntityConfig
from app.platform.crud_generator import build_pydantic_schema


def normalize(s: str) -> str:
    """Normalize a column name: lowercase, replace spaces and hyphens with underscores."""
    return s.lower().replace(" ", "_").replace("-", "_")


def auto_map_columns(
    file_columns: list[str],
    entity_fields: list[str],
) -> dict[str, str | None]:
    """Map file column names to entity field names using SequenceMatcher similarity.

    Uses a threshold of 0.6. Each entity field can only be matched once (greedy,
    first-best). Columns that do not meet the threshold are mapped to None.

    Args:
        file_columns: Column names from the uploaded file.
        entity_fields: Field names from the entity config.

    Returns:
        Dict mapping each file column to an entity field name or None.
    """
    available = list(entity_fields)
    mapping: dict[str, str | None] = {}

    for col in file_columns:
        col_norm = normalize(col)
        best_field: str | None = None
        best_score = 0.0

        for ef in available:
            ef_norm = normalize(ef)
            score = SequenceMatcher(None, col_norm, ef_norm).ratio()
            if score > best_score:
                best_score = score
                best_field = ef

        if best_score >= 0.6 and best_field is not None:
            mapping[col] = best_field
            available.remove(best_field)
        else:
            mapping[col] = None

    return mapping


def parse_csv(content: bytes, config: EntityConfig) -> dict:
    """Parse CSV file content and return columns, preview rows, all rows, and auto-mapping.

    Decodes with utf-8-sig to strip BOM. All cell values are strings.

    Args:
        content: Raw CSV file bytes from upload.
        config: EntityConfig for the target entity.

    Returns:
        Dict with keys: format, sheets, columns, preview_rows, all_rows, auto_mapping.
    """
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    columns = list(reader.fieldnames or [])
    rows = [{k: (str(v) if v is not None else "") for k, v in row.items()} for row in reader]

    entity_field_names = [f.name for f in config.fields]
    mapping = auto_map_columns(columns, entity_field_names)

    return {
        "format": "csv",
        "sheets": None,
        "columns": columns,
        "preview_rows": rows[:5],
        "all_rows": rows,
        "auto_mapping": mapping,
    }


def parse_excel(content: bytes, config: EntityConfig) -> dict:
    """Parse an Excel file and return sheet names for selection.

    Uses read_only=True and data_only=True to prevent formula evaluation
    and to avoid loading the full workbook into memory.

    Args:
        content: Raw Excel file bytes from upload.
        config: EntityConfig for the target entity (used for context, not parsing).

    Returns:
        Dict with keys: format, sheets, selected_sheet.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    return {
        "format": "excel",
        "sheets": sheet_names,
        "selected_sheet": sheet_names[0] if sheet_names else None,
    }


def parse_excel_sheet(content: bytes, sheet_name: str, config: EntityConfig) -> dict:
    """Parse a specific sheet from an Excel file.

    Extracts headers from row 0, preview from rows 1-3 (0-indexed), and all rows.
    All cell values are cast to str; None becomes "".

    Args:
        content: Raw Excel file bytes from upload.
        sheet_name: Name of the sheet to parse.
        config: EntityConfig for the target entity.

    Returns:
        Dict with keys: columns, preview_rows, all_rows, auto_mapping.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name]

    all_sheet_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_sheet_rows:
        return {
            "columns": [],
            "preview_rows": [],
            "all_rows": [],
            "auto_mapping": {},
        }

    # Row 0 = headers
    columns = [str(c) if c is not None else "" for c in all_sheet_rows[0]]

    # Rows 1+ = data
    data_rows: list[dict] = []
    for sheet_row in all_sheet_rows[1:]:
        row_dict = {
            columns[i]: (str(v) if v is not None else "")
            for i, v in enumerate(sheet_row)
            if i < len(columns)
        }
        data_rows.append(row_dict)

    entity_field_names = [f.name for f in config.fields]
    mapping = auto_map_columns(columns, entity_field_names)

    return {
        "columns": columns,
        "preview_rows": data_rows[:3],
        "all_rows": data_rows,
        "auto_mapping": mapping,
    }


def validate_rows(
    rows: list[dict],
    mapping: dict,
    config: EntityConfig,
) -> tuple[list[dict], list[dict]]:
    """Validate mapped rows against the entity's Pydantic schema.

    For each row, applies the column-to-field mapping, then validates via
    build_pydantic_schema. Returns separate lists of valid and error rows.

    Args:
        rows: List of dicts with file column names as keys.
        mapping: Dict mapping file column names to entity field names (or None).
        config: EntityConfig for the target entity.

    Returns:
        Tuple of (valid_rows, error_rows).
        valid_rows: List of dicts with entity field names as keys.
        error_rows: List of dicts with row_index, data, and error_reason keys.
    """
    ImportSchema = build_pydantic_schema(config, "ImportRow")

    valid_rows: list[dict] = []
    error_rows: list[dict] = []

    for i, row in enumerate(rows):
        # Apply mapping: translate file column names to entity field names
        mapped: dict[str, Any] = {}
        for file_col, entity_field in mapping.items():
            if entity_field is not None:
                mapped[entity_field] = row.get(file_col, "")

        try:
            validated = ImportSchema(**mapped)
            valid_rows.append(validated.model_dump())
        except Exception as e:
            error_rows.append({
                "row_index": i + 2,  # 1-based, row 1 is header
                "data": mapped,
                "error_reason": str(e),
            })

    return valid_rows, error_rows


def find_duplicates(
    rows: list[dict],
    config: EntityConfig,
    db_session: Any,
) -> list[dict]:
    """Detect rows that already exist in the database.

    Uses config.key_fields if non-empty; otherwise falls back to the first
    required string field. For each row, queries the DB for an existing record
    matching the key field values.

    Args:
        rows: List of dicts with entity field names as keys (post-mapping).
        config: EntityConfig for the target entity.
        db_session: SQLAlchemy session for DB queries.

    Returns:
        List of dicts with row_index, data, existing_id, key_field, key_value
        for each row that has a match in the database.
    """
    # Determine key fields
    if config.key_fields:
        key_field_names = config.key_fields
    else:
        # Fall back to first required string field
        fallback = next(
            (f.name for f in config.fields if f.required and f.type in ("string", "text")),
            None,
        )
        key_field_names = [fallback] if fallback else []

    if not key_field_names:
        return []

    # Use first key field for lookup (compound key support can be added later)
    key_field = key_field_names[0]

    duplicates: list[dict] = []
    for i, row in enumerate(rows):
        key_value = row.get(key_field)
        if key_value is None:
            continue

        # Dynamic query: look up any record where key_field column == key_value.
        # This requires the DB model to be discoverable; we use a raw text query
        # so the service remains decoupled from specific ORM models.
        from sqlalchemy import text
        result = db_session.execute(
            text(f"SELECT id FROM {config.table} WHERE {key_field} = :val LIMIT 1"),
            {"val": key_value},
        ).fetchone()

        if result is not None:
            duplicates.append({
                "row_index": i,
                "data": row,
                "existing_id": str(result[0]),
                "key_field": key_field,
                "key_value": key_value,
            })

    return duplicates


def generate_error_csv(error_rows: list[dict], config: EntityConfig) -> str:
    """Generate a CSV string containing error rows with an appended error_reason column.

    Args:
        error_rows: List of dicts with row_index, data, and error_reason keys.
        config: EntityConfig for the target entity (provides field names for column order).

    Returns:
        CSV string with entity field columns + error_reason column.
    """
    entity_field_names = [f.name for f in config.fields]
    fieldnames = entity_field_names + ["error_reason"]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for err in error_rows:
        row_data = dict(err.get("data", {}))
        row_data["error_reason"] = err.get("error_reason", "")
        writer.writerow(row_data)

    return output.getvalue()
